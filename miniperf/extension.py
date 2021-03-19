import logging
import os
import time
import traceback
import shutil
import sys
import threading
import datetime
import time
import glob
import wcwidth
import importlib
import subprocess
import re
import csv
from miniperf import setting
from miniperf import checker
from miniperf import custom_handler
from openpyxl import load_workbook
from miniperf import functionsEx
from miniperf import utils_str
ROOT_DIR = os.path.abspath(os.path.dirname(__file__))
python_path = os.path.join(os.path.split(sys.executable)[0], "python.exe")

g_webview = None


def registered_webview(webview):
    global g_webview
    g_webview = webview

def mkdir_if_not_exists(path, *paths):
    dirname = os.path.join(path, *paths)
    if not os.path.exists(dirname):
        os.mkdir(dirname)
    return dirname

def external_data_dir():
    return mkdir_if_not_exists(ROOT_DIR, ".data")

def post_message_to_js(msg, type):
    global g_webview
    if g_webview:
        g_webview.send_message_to_js({"type": "on_message", "data": {"type": type, "msg": msg}})

def rpc_get_file_list(xlsxpath):
    print("rpc_GetFileList")
    file_list = []
    try:
        files_status = get_svn_status(xlsxpath)
        print(files_status)

        files = glob.glob(xlsxpath + "/**/*.xls*", recursive=True)
        for f in files:
            if not os.path.basename(f).startswith("~"):
                status = files_status.get(f, "o")
                file_list.append({"name": os.path.basename(f), "path": f, "status": status})

        # list = output.replace(' ', '')
        # list = list[2:]
        # list = list.split(r'\n')
        # for f in list:
        #     name = f[1:-2].split(r'\\')[-1]
        #     if 'xls' in name and not name.startswith('~'):
        #         file_list.append({'name': name, 'path': f[1:-2].replace('\\\\','\\'), 'status':f[0]})
    except Exception as e:
        print(e)
        files = glob.glob(xlsxpath + "/**/*.xls*", recursive=True)
        for f in files:
            if not os.path.basename(f).startswith("~"):
                file_list.append({"name": os.path.basename(f), "path": f, "status": "o"})
    # print(xlsxpath)
    # files = glob.glob(xlsxpath + '/**/*.xls*', recursive=True)
    # for f in files:
    #     if not os.path.basename(f).startswith('~'):
    #             file_list.append({'name': os.path.basename(f), 'path': f})
    print(file_list)
    return file_list


def get_cstemplate_file():
    return os.path.join(ROOT_DIR, "asset", "template", "CSTemplate.cs")


def get_tpsvn():
    return os.path.join(ROOT_DIR, "asset", "tpsvn", "svn.exe")


def get_tpsvngui():
    return os.path.join(ROOT_DIR, "asset", "tpsvn", "TortoiseProc.exe")
    # return "C:\\Program Files\\TortoiseSVN\\bin\\TortoiseProc.exe"

def get_svn_status(xlsxpath):
    ret = {}
    output = subprocess.check_output([get_tpsvn(), "status", xlsxpath]).decode().replace("\r", "").replace(" ", "")
    lines = output.split("\n")
    for line in lines:
        status = line[:1]
        name = line[1:]
        # ret.append({'status': line[:1], 'name': line[1:]})
        if status == "?":
            status = "M"
        ret[name] = status
    return ret


def gen_cscode(xlsxpath:str):
    print('-------------------------------------------------gen_cscode')
    if not os.path.exists(xlsxpath):
        return {"ok": False, "msg": f"未找到对应文件/{xlsxpath}"}

    try:
        wb = load_workbook(filename=xlsxpath)
        ws = wb.active

        row_count = ws.max_row
        if row_count < 4:
            return {"ok": False, "msg": "配置表有异常,Row < 4"}

        tab_name = os.path.basename(xlsxpath).split(".")[0]

        comments = ws[1]

        #创建属性 和 自定义 Class
        fileds          = []
        filedNames      = {}
        classes         = []
        template_filed  = "\t\tpublic __FILED__; // __COMMENT__\n"
        functionsEx.csharp_custom_class.clear()#清空
        functionsEx.csharp_custom_enum.clear()
        for i in range(0, len(comments)):

            #[主类型1，主类型2，主属性名，第二层类型，第二层属性名,变量类型]
           
           #为空 跳过
            if ws[1][i].value == None or ws[5][i].value == None : 
                #print("关键值为空跳过",ws[1][i].value,ws[5][i].value)
                continue;

            results = functionsEx.analysis_Class(ws[1][i].value,ws[5][i].value)

            _type = results[5]
  
            if results[0] == "enum" :
               #枚举处理
               #('enum', 'ESex', 'sex', '', '', 'ESex') 
               args = utils_str.get_fun_args(ws[2][i].value,"EEnum")
               #['男-man-1','女-woman-2']
               #['humans-1','orc-2']

               for arg in args :

                    argsp = arg.split('-')
                    attributes = argsp
                    if len(argsp) == 3:
                        attributes = argsp[1:3]

                    functionsEx.add_custom_enum(results[1],attributes)

            else:
                #收集需要创建的 Class
                if not (results[0] in functionsEx.csharp_types) : 
                    functionsEx.add_custom_class(results[0],results[3],results[4],False)

                elif not (results[1] in functionsEx.csharp_types) :      
                    functionsEx.add_custom_class(results[1],results[3],results[4],True)   


            header = functionsEx.Header(
                _comment = ws[4][i].value,
                _tags    = ws[2][i].value,
                _filed   = results[2],#名字
                _type    = _type,
            )

            #为空 跳过
            # if header.Filed == 'None':
            #     continue

             #变量名字已经记录，无需在替换
            if results[2] in filedNames.keys() : continue
            filedNames[results[2]] = True    

            line = template_filed.replace("__FILED__", f"{header.FiledType} {header.Filed}")
            line = line.replace("__COMMENT__", f"{header.Comment}")
            fileds.append(line)
 
        #生成自定义的class     
        for item in functionsEx.csharp_custom_class.values() :
            class_str = functionsEx.gen_custom_class(item)
            print(class_str)
            classes.append(class_str)

        #生成自定义的enum
        for item in functionsEx.csharp_custom_enum.values() :
            enum_str = functionsEx.gen_custom_enum(item)
            print(enum_str)
            classes.append(enum_str)

        ########## 解析部分    
        header_list     = []
        fileds_parser   = []
        history         = {}#记录
        template_filed  = '\t\t\t__FILED__ = Get___TYPE__(cellStrs[__INDEX__], "");\n'
        template_enum   = '\t\t\t__FILED__ = Get_Enum<__T__>(cellStrs[__INDEX__]);\n'
        c_col = 0#列
        for i in range(0, len(comments)):
        #for start
            header = functionsEx.Header(
                _comment=ws[4][i].value,
                _tags=ws[2][i].value,
                _filed=ws[5][i].value,
                _type=ws[1][i].value,
            )
            header_list.append(header)
            if not header.Filed:
                return {"ok": False, "msg": f"[2]{i} is None"}

            if header.Filed == 'None':
                continue

            #[主类型1，主类型2，主属性名，第二层类型，第二层属性名,变量类型]
            results = functionsEx.analysis_Class(ws[1][i].value,ws[5][i].value)
            #results: ('List', 'DropItem', 'items1', 'string', 'name','List<DropItem>') 

            line = ""
            if results[0] == "enum" :
                #('enum', 'ESex', 'sex', '', '', 'ESex')                 --枚举类型
                line = template_enum.replace("__FILED__", header.Filed)
                line = line.replace("__T__",results[1])
                line = line.replace("__INDEX__", f"{c_col}")

            elif results[0] in functionsEx.csharp_simpleBase : 
                #简单类型    
                line = template_filed.replace("__FILED__", header.Filed)
                line = line.replace("__TYPE__", header.FiledType)
                line = line.replace("__INDEX__", f"{c_col}")
            else:
                #给复杂类型 instance
                if results[3] != "" :
                    get_sub_val = 'Get___TYPE__(cellStrs[__INDEX__],"")'.replace("__TYPE__",results[3])
                else:
                    get_sub_val = 'Get___TYPE__(cellStrs[__INDEX__],"")'.replace("__TYPE__",results[1])

                get_sub_val = get_sub_val.replace("__INDEX__",f"{c_col}")
                line   = functionsEx.instance_custom_class(results,history,get_sub_val,c_col)

            if line != "" :
                fileds_parser.append(line)

            c_col = c_col + 1
        #for end    

        #给复杂类型赋值
        #history
        add_parser = []
        print(history)
        exlines = functionsEx.set_custom_class(history)
        for line in exlines:
            add_parser.append(line)

        mark_pkey = []
        for e in header_list:
            if e.Tags and "PKEY" in e.Tags:
                k = e.Tags.split("|")[0]
                t = e.FiledType
                n = e.Filed
                mark_pkey.append((k, t, n))

        custom_code = "//----------Custom-Start----------\n\t\t//----------Custom-End----------"

        code_full_path = os.path.join(setting.cscode_dir, f"Tab{tab_name}.cs")
        if os.path.exists(code_full_path):
            with open(code_full_path, "r", encoding="utf-8") as f3:
                dat = f3.read()
                custom_re = re.compile(
                    r"//----------Custom-Start----------(?P<Custom>.*)//----------Custom-End----------", re.DOTALL
                )
                m = custom_re.findall(dat)
                if len(m) > 0:
                    custom_code = f"//----------Custom-Start----------{m[0]}//----------Custom-End----------"

        with open(get_cstemplate_file(), "r", encoding="utf-8") as f:
            dat = f.read()
            dat = dat.replace("<TABNAME>", tab_name)
            dat = dat.replace("<TABPATH>", f'"{tab_name}.csv"')
            dat = dat.replace("//MARK_FILEDS", "".join(fileds))
            dat = dat.replace("//MARK_CLASS_FILED", "".join(classes))
            dat = dat.replace("//MARK_PARSER", "".join(fileds_parser))
            dat = dat.replace("//MARK_ADD_METHOD", "".join(add_parser))
            
            for (k, t, n) in mark_pkey:
                dat = dat.replace(f"//MARK_T_{k}", t)
                dat = dat.replace(f"//MARK_N_{k}", n)

            if len(mark_pkey) == 1:
                dat = dat.replace(f"//1st_dec ", "")
                dat = dat.replace(f"//1st_add ", "")
                dat = dat.replace(f"//1st_get ", "")
            elif len(mark_pkey) == 2:
                dat = dat.replace(f"//2nd_dec ", "")
                dat = dat.replace(f"//2nd_add ", "")
                dat = dat.replace(f"//2nd_get ", "")

            dat = dat.replace("//Custom-Code", custom_code)

            with open(code_full_path, "wb") as f2:
                f2.write(dat.encode("utf-8"))
        return {"ok": True, "dat": f"【Succeed】生成成功\n{code_full_path}"}
    except Exception as err:
        print("Exception", err)
        traceback.print_exc()
        return {"ok": False, "msg": f"ERROR: {err}\n{traceback.format_exc()}"}



def gen_mgrcode(xlsxpath:str):
    mgrpath = setting.csmgr_path
    if not os.path.exists(xlsxpath) or not os.path.exists(mgrpath):
        return {"ok": False, "msg": f"未找到对应文件/{xlsxpath} or {mgrpath}"}

    try:
        tab_name = os.path.basename(xlsxpath).split(".")[0]
        tab_class_name = f"Table{tab_name}"
        tab_var_name = f"Table_{tab_name}"
        dat = ""
        with open(mgrpath, "r", encoding="utf-8") as f:
            dat = f.read()

        init_block = f"//{tab_class_name}-S-Init\n\t\t\t{tab_var_name} = new {tab_class_name}();\n\t\t\t{tab_var_name}.Init();\n\t\t\t//{tab_class_name}-E-Init\n\t\t\t//MARK_Init"
        async_block = f"{tab_var_name} = new {tab_class_name}();\n\t\t\tyield return {tab_var_name}.LoadAsync(TABLE_DIRECTORY);\n\t\t\t//MARK_Async"
        ondestory_block = f"//{tab_class_name}-S-OnDestory\n\t\t\t{tab_var_name} = null;\n\t\t\t//{tab_class_name}-E-OnDestory\n\t\t\t//MARK_OnDestory"
        filed_block = f"public {tab_class_name} {tab_var_name};\n\t\t//Mark_Fileds"
        if f"//{tab_class_name}-S-Init" not in dat:
            dat = dat.replace("//MARK_Init", init_block)
            dat = dat.replace("//MARK_Async", async_block)
            dat = dat.replace("//MARK_OnDestory", ondestory_block)
            dat = dat.replace("//Mark_Fileds", filed_block)

        with open(mgrpath, "wb") as f2:
            f2.write(dat.encode("utf-8"))
        return {"ok": True, "dat": f"【Succeed】生成成功\n{mgrpath}"}
    except Exception as err:
        print("Exception", err)
        traceback.print_exc()
        return {"ok": False, "msg": "应用产生异常,请联系开发人员"}


def dump(xlsxpath):
    def wstring(s, width):
        count = wcwidth.wcswidth(s) - len(s)
        width = width - count if width >= count else 0
        return "{0:{1}{2}{3}}".format(s, "", "^", width)

    wb = load_workbook(filename=xlsxpath)
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        cells = [wstring(f"{e.value}", 20).replace("\n", "") for e in ws[i]]
        print("|".join(cells))


def gen_tabfile(dat: dict):
    print('-----------------------------gen_tabfile')
    try:
        xlsxpath = dat["path"]
        tabdir = setting.table_dir
        if not os.path.exists(xlsxpath):
            return {"ok": False, "msg": f"未找到对应文件/{xlsxpath}"}
        if not os.path.exists(tabdir):
            return {"ok": False, "msg": f"未找到对应文件/{tabdir}"}

        wb = load_workbook(filename=xlsxpath)
        ws = wb.active
        row_count = ws.max_row
        if row_count < 4:
            return {"ok": False, "msg": "配置表有异常,Row < 4"}

        tab_name = os.path.basename(xlsxpath).split(".")[0]
        tab_fullpath = os.path.join(setting.table_dir, f"{tab_name}.csv")
        tab_headers = []
        for i in range(0, len(ws[1])):
            header = functionsEx.Header(
                _comment=ws[4][i].value,
                _tags=ws[2][i].value,
                _filed=ws[5][i].value,
                _type=ws[1][i].value,
            )
            tab_headers.append(header)

        # wb = openpyxl.load_workbook('test.xlsx')
        # sh = wb.active # was .get_active_sheet()
        # with open('test.csv', 'w', newline="") as f:
        #     c = csv.writer(f)
        #     for r in sh.rows:
        #         c.writerow([cell.value for cell in r])

        with open(tab_fullpath, "w", encoding="utf-8", newline="") as f:
            c = csv.writer(f)
            i = 0


            # 获取有效列,第5行中有列为空代表此列为注释列
            valid_cols_indexs = []
            for i, v in enumerate(ws[5]):
                if v.value != None:
                    valid_cols_indexs.append(i)            

            #获取第二行的，扩展操作标记        
            extend_closs = []        
            for i, v in enumerate(ws[2]):
                extend_closs.append(v.value)    
                   

            # 从第4行开始遍历
            irows = ws.iter_rows(min_row=4)
            taglist = {}
            rowCout = 4
            for r in irows:

                cols = [f"{cell.value}" for cell in r]

                for j in range(len(cols)):
                    if cols[j] == "None":
                        cols[j] = ""  # tab_header[j].Default

                    if "PATH" in tab_headers[j].Tags:
                        cols[j] = cols[j].replace("\\", "/")
                        cols[j] = cols[j].replace("\n", "")

                    if rowCout > 5 :     
                        extend = extend_closs[j]
                        if(extend != None):
                           
                            extend = extend.strip()#要剔除空格，请开启
                            ok,msg = custom_handler.custom(rowCout,j,extend,taglist,cols[j])
                            if ok :
                                cols[j] = msg
                            else:
                                return {"ok": False, "msg": f"ERROR:{msg}\n{traceback.format_exc()}"}   

                rowCout = rowCout + 1            
                # c.writerow([cell.value for cell in r])
                # for i in range(3, ws.max_row + 1):
                #     cells = [e.value for e in ws[i]]
                #     for j in range(len(cells)):
                #         if cells[j] == None:
                #             cells[j] = ''  # tab_header[j].Default
                #         if 'PATH' in tab_headers[j].Tags:
                #             cells[j] = cells[j].replace('\\', '/')

                #     # line = ','.join(cells)
                #     # if i != ws.max_row:
                #         # line += '\n'
                #     # f.write(line.encode("utf-8"))
                valid_cols = [cols[i] for i in valid_cols_indexs]

                c.writerow(valid_cols)
        return {"ok": True, "dat": f"【Succeed】生成成功\n{tab_fullpath}"}        
    except Exception as err:
        print("Exception", err)
        traceback.print_exc()
        return {"ok": False, "msg": f"ERROR: {err}\n{traceback.format_exc()}"}


def rpc_open_folder(dir):
    if ".cs" in dir:
        dir = os.path.dirname(dir)
    open_dir(dir)
    return {"ok": True, "dat": f"{dir}"}


def return_log():
    pass


def readOut(process):
    logging.info("readOut")
    for readOut_line in process.stdout:
        if len(readOut_line) < 1:
            break
        readOut_line = readOut_line.decode("utf-8")
        logging.info("readOut: %s", readOut_line)
        post_message_to_js(readOut_line, "out")


def readErr(process):
    logging.info("readErr")
    for readErr_line in process.stderr:
        if len(readErr_line) < 1:
            break
        readErr_line = readErr_line.decode("utf-8")
        logging.info("readErr: %s", readErr_line)
        post_message_to_js(readErr_line, "out")


def open_dir(report_dir):
    logging.info("open_dir" + report_dir)
    os.startfile(os.path.abspath(report_dir))


def checkfile(xlsxpath: str):
    ret, msg = (False, "【SUCCEED】通过通用性检查")
    try:
        wb = load_workbook(filename=xlsxpath)
        ws = wb.active

        ret, msg = checker.empty_check(ws)
        if ret == False:
            return {"ok": False, "msg": msg}

        ret, msg = checker.linecount_check(ws)
        if ret == False:
            return {"ok": False, "msg": msg}

        ret, msg = checker.pkey1_check(ws)
        if ret == False:
            return {"ok": False, "msg": msg}

        ret, msg = checker.tags_check(ws)
        if ret == False:
            return {"ok": False, "msg": msg}

        ret, msg = checker.custom_check(ws, xlsxpath)
        if ret == False:
            return {"ok": False, "msg": msg}

    except Exception as e:
        return {"ok": False, "msg": f"{e}"}

    return {"ok": True, "msg": "【SUCCEED】通过通用性检查"}



def rpc_OpenXlsxFileClick(dat):
    os.startfile(os.path.abspath(dat["path"]))


def rpc_CheckTableClick(dat):
    return checkfile(dat["path"])


def rpc_GenTabFileClick(dat):
    return gen_tabfile(dat)

def rpc_GenCSAndTabCodeClick(dat):
    ret = checkfile(dat['path'])
    if ret['ok'] == False:
        return ret
    ret = gen_cscode(dat['path'])
    if ret['ok'] == False:
        return ret
    ret = gen_mgrcode(dat['path'])
    if ret['ok'] == False:
        return ret
    ret = gen_tabfile(dat)
    if ret['ok'] == False:
        return ret

    return {'ok': True, 'dat': "【Succeed】生成对应CS代码和CSV配置文件成功"}

def rpc_CommitCSAndTabCodeClick(dat):
    p = "*".join([setting.cscode_dir, setting.table_dir, setting.xlsx_dir])
    try:
        c = get_tpsvngui()
        os.system(f'{c} /command:commit /path:"{p}"')
    except Exception as e:
        print(str(e))
    return {"ok": True, "dat": p}

def rpc_DefaultSetting(app):
    xlsx_dir = app.get_config("Settting/XLSX_PATH", os.path.join(".", "configs.ini"))
    setting.xlsx_dir = os.path.abspath(xlsx_dir)
    cscode_dir = app.get_config("Settting/CSCODE_PATH", os.path.join(".", "configs.ini"))
    setting.cscode_dir = os.path.abspath(cscode_dir)
    csmgr_path = app.get_config("Settting/MGR_PATH", os.path.join(".", "configs.ini"))
    setting.csmgr_path = os.path.abspath(csmgr_path)
    table_dir = app.get_config("Settting/TAB_PATH", os.path.join(".", "configs.ini"))
    setting.table_dir = os.path.abspath(table_dir)

    dat = {
        "XLSX_PATH": xlsx_dir,
        "CSCODE_PATH": cscode_dir,
        "TAB_PATH": table_dir,
        "MGR_PATH": csmgr_path,
    }
    return {"ok": True, "dat": dat}


def rpc_CommitFileClick(dat):
    pxlsx = os.path.abspath(dat["path"])
    pcsv = os.path.join(setting.table_dir, dat["name"].replace("xlsx", "csv"))
    try:
        pcsv_meta = pcsv + ".meta"
        c = get_tpsvngui()
        p = "*".join([pxlsx, pcsv, pcsv_meta])
        print(pxlsx, pcsv)
        os.system(f'{c} /command:commit /path:"{p}"')
    except Exception as e:
        print(str(e))
    return {"ok": True, "msg": pcsv}


def rpc_CommitAllFilesClick(dat):
    p = "*".join([setting.xlsx_dir, setting.table_dir])
    c = get_tpsvngui()
    os.system(f'{c} /command:commit /path:"{p}"')
    return {"ok": True, "msg": ""}


def rpc_ConvertAllFilesClick(dat):
    try:
        xlsxs = rpc_get_file_list(setting.xlsx_dir)
        xlsx_paths = [e["path"] for e in xlsxs]
        for p in xlsx_paths:
            dat = checkfile(p)
            if dat["ok"] == False:
                return dat
        for p in xlsx_paths:
            dat = gen_tabfile({"path": p})
            if dat["ok"] == False:
                return dat

        s_xlsx = "\n".join(xlsx_paths)
        return {"ok": True, "dat": f"【Succeed】转表成功\n{s_xlsx}"}
    except Exception as err:
        print("Exception", err)
        traceback.print_exc()
        return {"ok": False, "msg": f"ERROR: {err}\n{traceback.format_exc()}"}

def rpc_HelpButtonClick(dat):
    try:
        p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ui", "help.html")
        os.startfile(p)
        return {"ok": True, "dat": f"【Succeed】转表成功"}
    except Exception as err:
        print("Exception", err)
        traceback.print_exc()
        return {"ok": False, "msg": f"ERROR: {err}\n{traceback.format_exc()}"}