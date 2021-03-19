import importlib
import os
import re
from openpyxl import workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from miniperf import utils_str
def empty_check(ws):

    for i, row in enumerate(ws.rows):        
        if i == 2:
            continue
        line = [f'{e.value}' for e in row]
        if len(set(line)) == 1 and line[1] == 'None':
            return (False, "【FAILED】XLSX 有空行")
    #项目组需求，支持空列        
    # for row in ws.columns:
    #     line = [f'{e.value}' for e in row]
    #     if len(set(line)) == 1 and line[1] == 'None':
    #         return (False, "【FAILED】XLSX 有空列")
    return (True, "")

def linecount_check(ws:Worksheet):
    if ws.max_row < 4:
        return (False, "[ERROR1] 配置表有异常,Row < 4")
    return (True, "")

def pkey1_check(ws:Worksheet):
    tags = [f'{e.value}' for e in ws[2]]
    if 'PKEY1' not in ''.join(tags):
        return (False, f"[ERROR2] 配置表需要要有主键PKEY1, 请联系开发人员 {tags}")
    return (True, "")

                  
def tags_check(ws:Worksheet):
    s_cols = [f'{e.value}' for e in ws[2]]
    taglist = []
    for i in range(len(s_cols)):
        if s_cols[i] != 'None':
            tag_list = s_cols[i].split('|')
            for tag in tag_list:
                func, args = utils_str.tag_parser(tag)
                taglist.append((i, func, args))

    for i, func, args in taglist:
        if func[:1] == "F":
            modulename = func
            spec = importlib.util.find_spec(modulename)
            if spec:
                module = importlib.reload(importlib.__import__(modulename))
                if module:
                    cletter = get_column_letter(i+1)
                    cols = [f'{e.value}' for e in ws[cletter][4:]]
                    # i 代表列
                    ok, msg = module.check(cols, args,i)
                    if not ok:
                        return (False, f"【ERROR】 {msg}")
    return (True, "【SUCCEED】 标签检查")

def ws_to_dict(ws:Worksheet):
    wsdict = {}
    for citer in ws.iter_cols():
        v = citer[3].value
        name = f'{v}'
        if name != 'None':
            wsdict[name] =  [f'{e.value}' for e in citer[4:]]
            print(name, wsdict[name])
    return wsdict


def custom_check(ws:Worksheet, p:str):
    wsdict = ws_to_dict(ws)
    modulename = os.path.basename(p).split(".")[0]
    spec = importlib.util.find_spec(modulename)
    if spec:
        module = importlib.reload(importlib.__import__(modulename))
        if module:
            ok, msg = module.check(ws, wsdict)
            return (ok, msg)
    else:
        print(f'[INFO] Not Found {modulename}.py')

    return (True, "【SUCCEED】 自定义检查")
