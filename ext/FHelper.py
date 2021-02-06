import os
from openpyxl import load_workbook

app = None

def get_table(table_file_name:str):
    table = {}
    xlsx_dir = app.get_config("Settting/XLSX_PATH",'')
    xlsx_path = os.path.join(xlsx_dir, f'{table_file_name}.xlsx')
    if not os.path.exists(xlsx_path):
        return (False, table, f'{xlsx_path} 不存在')

    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    for i in range(0, len(ws[1])):
        col    = chr(ord('A')+i)
        x_cols = [f'{e.value}' for e in ws[col][4:]]
        name   = ws[col + '4'].value
        table[name] = x_cols

    return (True, table, '')